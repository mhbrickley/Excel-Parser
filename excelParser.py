'''
Excel Parser
This program parses tabs within an Excel workbook into separate files
based on user defined file extension and delimiter
'''

#import modules
import os
import openpyxl as op
import csv

#determine path
while True:
    try:
        #user input
        path = input('Enter the path of the folder containing the Excel file you wish to parse: ').strip()
        # change directory to user specified path
        os.chdir(path)
        break
    except FileNotFoundError:
        print('The path specified does not exist')
        continue

#remove existing parsed files
extension = input('Enter the extension for each newly parsed file (ex. \'.txt\'): ').strip()
for i in os.listdir(path):
    if i.endswith(extension):
        os.remove(i)

#create files
while True:
    try:
        delim = input('Enter the delimiter for each file [comma, pipe, etc.]: ').strip()
        #create CSV files
        #return all files within specified path
        for i in os.listdir(path):
            if i.endswith('.xlsx'):
                #initialize workbook
                wb = op.load_workbook(i,data_only=True)
                #get all sheets within workbook
                for i in wb.sheetnames:
                    #initialize CSV file
                    csvGen = open(i + '.csv','w')
                    #write to CSV file
                    csvFile = csv.writer(csvGen,delimiter=delim)

                    #create list of tabs
                    tabs = wb[i]
                    rowCount = tabs.max_row + 1
                    colCount = tabs.max_column + 1

                    #create rows
                    for i in range(1,rowCount):
                        #empty matrix to append data
                        xlData = []
                        #create columns
                        for j in range(1,colCount):
                            #find cell data
                            cellData = tabs.cell(row=i,column=j).value
                            #append cell data to empty matrix
                            xlData.append(cellData)
                        #append matrix data to CSV file
                        csvFile.writerow(xlData)
                    #close CSV file initializer
                    csvGen.close()

        #change .csv to user specified extension
        for i in os.listdir(path):
            if i.endswith('.csv'):
                #split pathname into root & extension
                root = os.path.splitext(i)[0]
                #rename file
                os.rename(i,root + extension)
        break

    except TypeError:
        print('Delimiter not valid')
        continue

#completion notification
print('Parsing to',extension,'complete')